"""
测试用例 Excel 导出工具

提供将测试用例导出为 Excel 文件的能力，支持企业级测试管理工具的导入格式。
"""

import io
import json
from pathlib import Path
from typing import Any

from langchain.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from app.config.settings import settings
from app.agents.tools.testcase.export_common import (
    EXPORT_HEADERS,
    case_key,
    extract_field,
    flatten_expected_results,
    flatten_preconditions,
    flatten_steps,
    flatten_test_data,
    localize_case_type,
    localize_priority,
)
from app.agents.tools.testcase.export_formats import (
    generate_test_cases_csv_bytes,
    generate_test_cases_json_bytes,
)
# pragma: no cover  MC80OmFIVnBZMlhsdEpUbXRiZm92b2s2UWtSWGRBPT06OWM0ZDYxMTc=

# 与 agent.py 中 composite_backend 的 "/" 路由保持一致：
# Agent 在虚拟文件系统中看到的路径以 "/" 为根，实际落盘到 workspace_root。
_WORKSPACE_ROOT = Path(settings.testcase_workspace_root).resolve()

_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ALIGNMENT_WRAP = Alignment(vertical="top", wrap_text=True)
_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_DEFAULT_COLUMN_WIDTHS = {
    "A": 18,
    "B": 35,
    "C": 14,
    "D": 12,
    "E": 10,
    "F": 30,
    "G": 40,
    "H": 30,
    "I": 40,
    "J": 20,
}



def _ensure_xlsx_suffix(path: Path) -> Path:
    """确保导出路径以 .xlsx 结尾，缺失时自动补全。"""
    if path.suffix.lower() != ".xlsx":
        return path.with_suffix(".xlsx")
    return path


def generate_test_cases_excel_bytes(
    test_cases: list[dict[str, Any]],
    sheet_name: str = "测试用例",
) -> bytes:
    """将测试用例列表转换为 Excel 文件字节流。

    字段兼容 export_test_cases_to_excel 中声明的别名体系。
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表")
    ws.title = sheet_name

    headers = EXPORT_HEADERS
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGNMENT_CENTER
        cell.border = _BORDER

    for case in test_cases:
        steps = extract_field(case, "steps", "测试步骤", "test_case_steps", "test_steps", default=None)
        expected_results = extract_field(case, "expected_results", "预期结果", "expected_result", "expected", default=None)
        # 如果 case 级别没有预期结果，但步骤里带 result，则自动聚合
        if not expected_results and steps:
            expected_results = extract_expected_results_from_steps(steps)

        row = [
            extract_field(case, "id", "用例编号", "identifier", "case_id", "case_number", "编号"),
            extract_field(case, "title", "用例标题", "name", "标题", "用例名称"),
            extract_field(case, "module", "所属模块", "module_name", "功能模块", "模块"),
            localize_case_type(extract_field(case, "type", "用例类型", "case_type", "测试类型", "类型")),
            localize_priority(extract_field(case, "priority", "优先级")),
            flatten_preconditions(extract_field(case, "preconditions", "前置条件", "precondition", default=None)),
            flatten_steps(steps),
            flatten_test_data(extract_field(case, "test_data", "测试数据", "data", default=None)),
            flatten_expected_results(expected_results),
            extract_field(case, "remarks", "备注", "remark", "note", "description", "desc", "描述"),
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _ALIGNMENT_WRAP
            cell.border = _BORDER

    for col_letter, width in _DEFAULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _to_virtual_path(real_path: Path) -> str:
    """把 workspace_root 下的真实路径反向映射为 Agent 可识别的虚拟路径。

    与 composite_backend 的虚拟文件系统一致：以 "/" 为根、使用正斜杠。
    """
    rel = real_path.resolve().relative_to(_WORKSPACE_ROOT)
    return "/" + rel.as_posix()


def _resolve_workspace_path(output_path: str | Path) -> Path:
    """将 Agent 传入的（虚拟）路径映射到真实 workspace_root 下。

    与 composite_backend 的 "/" 路由保持一致：
      - 绝对/虚拟根路径（如 "/测试用例.xlsx"）按相对 workspace_root 处理；
      - 相对路径（如 "测试用例.xlsx"）也落到 workspace_root 下；
      - 已经位于 workspace_root 内的真实绝对路径保持不变；
      - 缺少 .xlsx 后缀时自动补全。
    并禁止通过 ".." 越权写到 workspace_root 之外。
    """
    raw = Path(output_path)

    # 已经是 workspace_root 内的真实绝对路径，直接使用
    if raw.anchor:
        try:
            if raw.is_absolute() and raw.resolve().is_relative_to(_WORKSPACE_ROOT):
                return _ensure_xlsx_suffix(raw.resolve())
        except (ValueError, OSError):
            pass
        # 其余带锚点的路径（虚拟根 "/xxx"、盘符根等）剥离锚点后按相对处理
        anchor_len = len(Path(raw.anchor).parts)
        rel = Path(*raw.parts[anchor_len:]) if len(raw.parts) > anchor_len else Path()
    else:
        rel = raw

    if not rel.parts:
        raise ValueError(f"导出路径无效：{output_path}")

    resolved = (_WORKSPACE_ROOT / rel).resolve()
    if not resolved.is_relative_to(_WORKSPACE_ROOT):
        raise ValueError(
            f"导出路径越权：{output_path} 解析后超出工作目录 {_WORKSPACE_ROOT}"
        )
    return _ensure_xlsx_suffix(resolved)


def _resolve_input_path(input_path: str | Path) -> Path:
    """将 Agent 传入的（虚拟）输入文件路径映射到真实 workspace_root 下。

    与 _resolve_workspace_path 的映射规则一致（虚拟根/相对/真实绝对路径都落到
    workspace_root），但**不**强制 .xlsx 后缀，用于读取用例数据文件（.jsonl/.json）。
    同样禁止通过 ".." 越权读取 workspace_root 之外的文件。
    """
    raw = Path(input_path)

    if raw.anchor:
        try:
            if raw.is_absolute() and raw.resolve().is_relative_to(_WORKSPACE_ROOT):
                return raw.resolve()
        except (ValueError, OSError):
            pass
        anchor_len = len(Path(raw.anchor).parts)
        rel = Path(*raw.parts[anchor_len:]) if len(raw.parts) > anchor_len else Path()
    else:
        rel = raw

    if not rel.parts:
        raise ValueError(f"输入文件路径无效：{input_path}")

    resolved = (_WORKSPACE_ROOT / rel).resolve()
    if not resolved.is_relative_to(_WORKSPACE_ROOT):
        raise ValueError(
            f"输入文件路径越权：{input_path} 解析后超出工作目录 {_WORKSPACE_ROOT}"
        )
    return resolved


def _parse_json_objects(text: str, source: str) -> list[Any]:
    """从文本中解析出所有 JSON 对象，对常见的「脏」格式有强容错。

    无需调用方保证文件是严格的 JSONL，自动兼容：
      - 标准 JSONL：每行一个 JSON 对象
      - 整文件 JSON 数组：[ {...}, {...} ]
      - 多个对象挤在同一行、对象跨多行、对象之间用逗号或空白分隔
        （例如 LLM 直接拼接产出的 `{...},{...},{...}`）

    实现上用 json.JSONDecoder().raw_decode 顺序扫描，不依赖换行边界，
    因此 Agent 不再需要在 LLM 侧手工合并/规整文件，从根本上规避
    「跨行 JSON 解析失败」「合并时 token 截断」等问题。
    """
    text = text.strip()
    if not text:
        return []

    # 整文件 JSON 数组：直接解析
    if text.startswith("["):
        data = json.loads(text)
        if not isinstance(data, list):
            raise ValueError(f"用例数据文件顶层不是 JSON 数组：{source}")
        return data

    # 其余情况：顺序扫描出连续的 JSON 值，容忍逗号/空白分隔
    decoder = json.JSONDecoder()
    objs: list[Any] = []
    idx, n = 0, len(text)
    while idx < n:
        # 跳过对象之间的空白与逗号分隔符
        while idx < n and text[idx] in " \t\r\n,":
            idx += 1
        if idx >= n:
            break
        try:
            obj, end = decoder.raw_decode(text, idx)
        except json.JSONDecodeError as e:
            snippet = text[idx:idx + 80].replace("\n", " ")
            raise ValueError(
                f"用例数据文件 {source} 第 {text.count(chr(10), 0, idx) + 1} 行附近不是合法 JSON："
                f"{e}（片段：{snippet!r}）"
            ) from e
        objs.append(obj)
        idx = end
    return objs


def _describe_missing_file(requested: str | Path, resolved: Path) -> str:
    """构造「文件不存在」的自纠错提示。

    模型常传入宿主机绝对路径（如 D:\\sorted_cases.jsonl）或拼错文件名，
    仅回显原始路径无法让模型知道工具实际去哪找、工作目录里到底有哪些文件。
    这里同时给出解析后的真实路径，并列出工作目录下现有的 .jsonl/.json 数据文件，
    让模型据此改用正确的（虚拟）相对路径重试。
    """
    try:
        available = sorted(
            _to_virtual_path(f)
            for f in _WORKSPACE_ROOT.glob("*")
            if f.is_file()
            and f.suffix.lower() in (".jsonl", ".json")
            and f.name.lower() not in ("package.json", "package-lock.json")
        )
    except OSError:
        available = []
    avail_hint = "、".join(available) if available else "（工作目录下暂无 .jsonl/.json 数据文件）"
    return (
        f"用例数据文件不存在：{requested}\n"
        f"工具实际查找的真实路径：{resolved}\n"
        f"提示：input_file 应使用工作目录下的（虚拟）相对路径，不要用 D:\\ 等宿主机绝对路径。\n"
        f"当前工作目录下可用的数据文件：{avail_hint}"
    )


def _load_test_cases_from_file(
    input_path: str | Path | list[str | Path],
    dedup: bool = True,
) -> list[dict[str, Any]]:
    """从工作目录下的一个或多个数据文件读取并合并测试用例列表。

    支持的格式（每个文件自动识别，详见 _parse_json_objects）：
      - JSONL：每行一个 JSON 对象（推荐，便于 Agent 分批追加写入，永不超 token 上限）
      - JSON ：整个文件是一个 JSON 数组 [ {...}, {...} ]
      - 容错：多个对象同行/跨行/逗号分隔的「脏」拼接格式

    input_path 可以是单个路径，也可以是路径列表——传列表时按顺序读取并合并所有文件，
    Agent 无需先把多个文件手工合并成一个，直接把文件清单交给本工具即可。

    dedup=True 时按用例编号（或标题）去重，后出现的覆盖先出现的，
    用于消除多文件之间的重复用例。
    """
    paths = input_path if isinstance(input_path, (list, tuple)) else [input_path]

    cases: list[dict[str, Any]] = []
    for p in paths:
        real_path = _resolve_input_path(p)
        if not real_path.is_file():
            raise FileNotFoundError(_describe_missing_file(p, real_path))
        text = real_path.read_text(encoding="utf-8").strip()
        if not text:
            # 多文件场景下允许个别文件为空，跳过即可；单文件为空仍然报错
            if len(paths) == 1:
                raise ValueError(f"用例数据文件为空：{p}")
            continue
        cases.extend(_parse_json_objects(text, str(p)))

    invalid = [i for i, c in enumerate(cases) if not isinstance(c, dict)]
    if invalid:
        raise ValueError(
            f"用例数据文件存在非对象元素（下标 {invalid[:5]}...），每条用例必须是 JSON 对象。"
        )

    if dedup:
        merged: dict[Any, dict[str, Any]] = {}
        no_key: list[dict[str, Any]] = []
        for case in cases:
            key = case_key(case)
            if key is None:
                no_key.append(case)
            else:
                merged[key] = case  # 后出现的覆盖先出现的
        cases = list(merged.values()) + no_key

    return cases


@tool
def export_test_cases_to_excel(
    output_path: str | Path,
    test_cases: list[dict[str, Any]] | None = None,
    input_file: str | Path | list[str | Path] | None = None,
    sheet_name: str = "测试用例",
    dedup: bool = True,
) -> str:
    """
    将测试用例导出为 Excel 文件。

    用例来源二选一（推荐使用 input_file 以规避「用例多时数据截断」）：
      - input_file: 工作目录下的用例数据文件路径（.jsonl 或 .json），**可传单个路径，
        也可传路径列表一次性合并多个文件**。**用例数量多或分散在多个文件时优先用此方式**：
        先用文件写入工具把用例追加进 .jsonl 文件，再把文件（或文件清单）交给本工具读取导出。
        无需在对话里手工合并多个文件——直接传 input_file=["a.jsonl", "b.jsonl", ...] 即可，
        合并与去重在服务端完成，不受单次模型输出长度限制，永不截断。
        每个文件的格式自动识别并强容错：标准 JSONL（每行一对象）、整文件 JSON 数组、
        以及多个对象同行/跨行/逗号分隔的「脏」拼接格式都能正确解析。
      - test_cases: 直接内联传入的用例字典列表。仅适用于用例较少（约 < 30 条）的场景；
        用例过多会因模型单次输出 token 上限导致 JSON 被截断、数据丢失。
    两者同时提供时，input_file 优先。

    支持的测试用例字段（兼容 JSON / CSV / Markdown 中定义的格式）：
      - id / 用例编号
      - title / 用例标题
      - module / 所属模块
      - type / 用例类型
      - priority / 优先级
      - preconditions / 前置条件
      - steps / 测试步骤
      - test_data / 测试数据
      - expected_results / 预期结果
      - remarks / 备注

    Args:
        output_path: 导出文件路径。相对/虚拟根路径（如 "测试用例.xlsx" 或
            "/测试用例.xlsx"）会映射到 Agent 工作目录（workspace_root）下，
            与其他文件工具看到的虚拟文件系统保持一致。缺少 .xlsx 后缀时自动补全。
        test_cases: 测试用例字典列表，每个字典描述一条用例（用例少时使用）。
        input_file: 用例数据文件（.jsonl/.json）路径，或多个文件路径组成的列表（用例多时推荐）。
            映射规则同 output_path。传列表时按顺序读取并合并所有文件。
        sheet_name: 工作表名称，默认为 "测试用例"。
        dedup: 是否按用例编号（或标题）去重，默认 True；多文件合并时可消除重复用例。

    Returns:
        导出文件的虚拟路径字符串（以 "/" 为根），可直接交给其他文件工具复用。
    """
    if input_file is not None:
        test_cases = _load_test_cases_from_file(input_file, dedup=dedup)

    if not test_cases:
        raise ValueError(
            "没有可导出的测试用例：请通过 input_file 提供用例数据文件，或通过 test_cases 内联传入用例列表。"
        )

    output_path = _resolve_workspace_path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    excel_bytes = generate_test_cases_excel_bytes(test_cases, sheet_name)
    output_path.write_bytes(excel_bytes)
    return _to_virtual_path(output_path)
